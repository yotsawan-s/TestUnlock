#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ตัวอย่างการใช้งาน Excel Password Unlocker
"""

# ตัวอย่างที่ 1: ปลดล็อค Sheet Protection
def example_unlock_sheet():
    """ตัวอย่างการปลดล็อค Sheet Protection"""
    from openpyxl import load_workbook
    
    input_file = "locked_sheet.xlsx"
    output_file = "unlocked_sheet.xlsx"
    
    # โหลดไฟล์
    wb = load_workbook(input_file)
    
    # ปลดล็อคทุก sheet
    for sheet in wb.worksheets:
        sheet.protection.sheet = False
        sheet.protection.password = None
    
    # บันทึกไฟล์
    wb.save(output_file)
    print(f"✅ ปลดล็อคเรียบร้อย: {output_file}")


# ตัวอย่างที่ 2: ปลดล็อค File Protection (ต้องรู้ password)
def example_unlock_file():
    """ตัวอย่างการปลดล็อค File Protection"""
    import msoffcrypto
    
    input_file = "locked_file.xlsx"
    output_file = "unlocked_file.xlsx"
    password = "your_password_here"
    
    with open(input_file, "rb") as f:
        file = msoffcrypto.OfficeFile(f)
        file.load_key(password=password)
        
        with open(output_file, "wb") as out:
            file.decrypt(out)
    
    print(f"✅ ปลดล็อคเรียบร้อย: {output_file}")


# ตัวอย่างที่ 3: ปลดล็อคทั้งสองแบบ
def example_unlock_both():
    """ตัวอย่างการปลดล็อคทั้ง File และ Sheet Protection"""
    import msoffcrypto
    from openpyxl import load_workbook
    import os
    
    input_file = "locked_both.xlsx"
    temp_file = "temp.xlsx"
    output_file = "unlocked_both.xlsx"
    password = "your_password_here"
    
    # ขั้นตอนที่ 1: ปลดล็อค File Protection
    with open(input_file, "rb") as f:
        file = msoffcrypto.OfficeFile(f)
        file.load_key(password=password)
        
        with open(temp_file, "wb") as out:
            file.decrypt(out)
    
    print("✅ ปลดล็อค File Protection แล้ว")
    
    # ขั้นตอนที่ 2: ปลดล็อค Sheet Protection
    wb = load_workbook(temp_file)
    
    for sheet in wb.worksheets:
        sheet.protection.sheet = False
        sheet.protection.password = None
    
    wb.save(output_file)
    
    # ลบไฟล์ชั่วคราว
    os.remove(temp_file)
    
    print(f"✅ ปลดล็อคทั้งหมดเรียบร้อย: {output_file}")


if __name__ == "__main__":
    print("=" * 60)
    print("ตัวอย่างการใช้งาน Excel Password Unlocker")
    print("=" * 60)
    print()
    print("📝 หมายเหตุ: ไฟล์ตัวอย่างเหล่านี้เป็นเพียงโค้ดสาธิต")
    print("   ในการใช้งานจริง ให้ใช้ unlock_excel.py แทน")
    print()
    print("ตัวอย่างที่มี:")
    print("1. example_unlock_sheet() - ปลดล็อค Sheet Protection")
    print("2. example_unlock_file() - ปลดล็อค File Protection")
    print("3. example_unlock_both() - ปลดล็อคทั้งสองแบบ")
